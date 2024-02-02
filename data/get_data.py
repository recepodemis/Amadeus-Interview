import openpyxl

from constants.global_constants import *

def get_data():
    csvFile = openpyxl.load_workbook(CITIES_PATH)
    sheet = csvFile["Sheet1"]
    rows = sheet.max_row
    data = []
    for i in range(1,rows + 1):
        city = sheet.cell(i,1).value
        data.append(city)
    return data